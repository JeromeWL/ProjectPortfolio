import openpyxl
import datetime

#Opens the workbook and gets the most recent workbook save and saves it to the worksheet
try:
    wb = openpyxl.load_workbook('Expense tracker.xlsx')
    ws = wb.active
except FileNotFoundError as e:
     print(f'{e} file not found, check the name or path of the file')

#Gets empty cell value
def getEmptyCell():
    #Saves empty cell value
    emptyCell = ''
    
    #Loops through the active ws in column A
    for cell in ws['A']:
        #Check to see if there's gaps in the column
        if cell.value is None:
            emptyCell = cell.row
            break
        
    #If none found it will use a new line
    else:
        emptyCell = cell.row+1
    
    #Returns the emptyCell value
    return emptyCell

#Finds the column value of the wanted title
def findCol(str):
    #Loops through the cells in the first row
    for cell in ws['1']: 
        #Checks to make sure that the column title and the str input are the same
        if cell.value.lower() == str.lower():
            return cell.column
        
        #If no column is found
        if cell.value is None and ws.cell(row=cell.row, column=cell.column+1).value is None:
            print('Error: Title not found')
            #ends the function to prevent stack overflow
            return

#Calculates the running total on the excel spreadsheet
def calculateRTotal(emptyRow, amountPaid):
    amountCol = findCol('Amount paid')
    runTot = 0
    #If it is the first row it will not try to find the previous total becasue there isn't one. (Fence post statement)
    if emptyRow == 2:
        runTot = amountPaid
        
    #Otherwise it'll grab the previous total and add it to the current ammount paid
    else:
        runTot = amountPaid + (ws.cell(row=emptyRow-1, column=amountCol).value)
    
    #Updates the value
    ws.cell(row=emptyRow, column=findCol('Running total')).value = runTot


#Can filter your expenses to find how you much spent on a date or spent on gas
#category is what column you want to filter in
#value is the value you want to get the money spent
def filterBy(category, value):
    #Variable preperation
    amountCol = findCol('Amount paid')
    ToECol = findCol('Type of expense')
    spent = 0
    
    #Iterating through the categories
    for entry in ws.iter_cols(max_col=ToECol, max_row=1):
        for cell in entry:
            #Check to make sure the category is the same as the requested
            if cell.value.lower() == category.lower():
                #Iterates through the category's rows
                for values in ws.iter_rows(min_row=2, min_col=cell.column, max_col=cell.column):
                    for cellValues in values:
                        #Checks the specified value input by the user
                        if cellValues.value == value:
                            #Adds to the total spent
                            spent += ws.cell(row=cellValues.row, column=amountCol).value

    return spent
                    
#date in the format date datetime.datetime(year, month, day)
#method of payment card cash credit, etc..
#paidTo company etc..
#ammountPaid
def addExpense(date, methodOfPayment, typeOfExpense, amountPaid):
    #Finds the first empty row
    emptyRow = getEmptyCell()
    #More compressed code instead of calling each value to be chagned
    values = [date, methodOfPayment, typeOfExpense, amountPaid]
    
    #Loops through the values list and changes the value of the column accordingly
    for col, val in enumerate(values, start=1):
        ws.cell(row=emptyRow, column=col).value = val    
    
    #Passes the empty row value and the ammount paid to calculate the running total
    calculateRTotal(emptyRow, amountPaid)
    
    #Saves changes
    wb.save('Expense tracker.xlsx')

#addExpense(datetime.date(2004,6,15), 'Cash', 'Cosco', 12.40)
#addExpense(datetime.date(2004,2,12), 'Card', 'Kwik Trip', 104.10)
#addExpense(datetime.date(2004,2,11), 'Card', 'Target', 194.60)

#print(filterByMethod("Card"))
print(filterBy('Method of Payment', 'Card'))
print(filterBy('Type of expense', 'Cosco'))


#Functions to add
#Make a function to filter to get ammounts spent at a company, type of payment etc.